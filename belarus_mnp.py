import csv
import os
import shutil
import traceback
from datetime import datetime
from glob import glob
from pathlib import Path

import openpyxl

import exceptions
import utils
from utils.loger_util import create_logger


def parse_file(file_in, bel_settings):
    try:
        wb = openpyxl.load_workbook(file_in)
        sheet = wb['Sheet1']
        rows = []
        for row in sheet.iter_rows(0, sheet.max_row):
            data = []
            for cell in row:
                data.append(cell.value)
            rows.append(data)
        del rows[0]
        for row in rows:
            if not all(row):
                raise exceptions.BelarusParserError(
                    f'empty row found!\n\n{row}') from None
            row[0] = f'2570{row[0]}'
            row[2] = int(datetime.strptime(row[2],
                                           '%d.%m.%Y %H:%M:%S').timestamp())
            row[0], row[1] = row[1], row[0]
        with open(bel_settings.lock_file, 'w') as lock_mnp_db:
            csv_f = csv.writer(lock_mnp_db, delimiter=';')
            csv_f.writerows(rows)
    except Exception as err:
        raise exceptions.BelarusParserError(err)


def delete_tmp_files(settings: utils.BelSettings, *args):
    for file in args:
        if isinstance(file, Path):
            if file.exists():
                os.remove(file)
    if settings.lock_file.exists():
        os.remove(settings.lock_file)


def processing_mnp(base_settings: utils.BaseSettings,
                   bel_settings: utils.BelSettings):
    logger = create_logger(__name__, base_settings.log_dir)

    logger.info('start processing')
    source_file_mask = bel_settings.source_dir.joinpath(
        bel_settings.source_file_mask)
    source_files = glob(str(source_file_mask))

    if not source_files:
        msg = f'did not find files matching pattern {source_file_mask}'
        logger.warning(msg)
        utils.send_email(text=msg,
                         subject='Belarus mnp parser error')
        return
    elif len(source_files) != 1:
        msg = f'to many source files found in {bel_settings.source_dir}: {source_files}'
        logger.warning(msg)
        utils.send_email(msg, 'Belarus mnp parser error')
        return

    source_file = Path(source_files[0])

    if not os.stat(source_file).st_size:
        msg = f'{source_file=} is empty'
        logger.warning(msg)
        utils.send_email(msg, 'Belarus mnp parser error')
        return
    try:
        tmp_file = Path(shutil.copy(source_file, base_settings.tmp_dir))
        archive_path = utils.archive_file(source_file, bel_settings.archive_dir)

        if bel_settings.handled_file_path.exists():
            utils.archive_file(bel_settings.handled_file_path,
                               bel_settings.archive_dir)

        try:
            logger.info('start parse file')
            parse_file(tmp_file, bel_settings)
        except exceptions.ParserError as err:
            logger.exception(f'an exception during parsing \n\n{err}\n{archive_path=}', exc_info=True, stack_info=True)
            delete_tmp_files(bel_settings, source_file, tmp_file)
            tb = traceback.format_exc()
            utils.send_email(text=f'{err}\n\n{tb})',
                             subject='Belarus mnp parser error')
            return
        logger.info('finished parse file')
        shutil.move(bel_settings.lock_file, bel_settings.handled_file_path)
        delete_tmp_files(bel_settings, source_file, tmp_file)
        logger.info('finished processing')
    except Exception as err:
        logger.exception(msg='an error during processing\n\n{err}', exc_info=True, stack_info=True)
        delete_tmp_files(bel_settings, source_file, tmp_file)
        raise exceptions.MnpProcessingError(
            f"an error during processing Belarus's mnp\n\n{err}") from None
